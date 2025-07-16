# SPDX-License-Identifier: MIT
# SPDX-FileCopyrightText: 2024 Projected Journal

"""
Excel export functionality for the Projected Journal package.

This module provides utilities for writing DataFrames to Excel files with
proper formatting and styling.
"""

from pathlib import Path
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from .logging import get_logger

logger = get_logger(__name__)


def write_excel(df: pd.DataFrame, path: Path, sheet_name: str = "Journal") -> None:
    """
    Write a DataFrame to an Excel file with proper formatting.
    
    Features:
    - Freeze header row
    - Auto-fit column widths
    - Apply green table style
    - Format as Excel table
    
    Args:
        df: DataFrame to export
        path: Output file path
        sheet_name: Name of the Excel sheet (default: "Journal")
        
    Raises:
        OSError: If the file cannot be written
        ValueError: If DataFrame is empty
    """
    if df.empty:
        raise ValueError("Cannot export empty DataFrame")
    
    logger.info(f"Exporting {len(df)} rows to {path}")
    
    # Ensure output directory exists
    path.parent.mkdir(parents=True, exist_ok=True)
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write data to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Auto-fit column widths
    _autofit_columns(ws, df)
    
    # Format header row
    _format_header_row(ws)
    
    # Create Excel table with green style
    _create_table(ws, df, sheet_name)
    
    # Save workbook
    wb.save(path)
    logger.info(f"Successfully exported to {path}")


def _autofit_columns(ws, df: pd.DataFrame) -> None:
    """Auto-fit column widths based on content."""
    for idx, column in enumerate(df.columns, 1):
        column_letter = get_column_letter(idx)
        
        # Calculate max width needed
        # Check header length
        max_length = len(str(column))
        
        # Check data length (sample first 100 rows for performance)
        sample_data = df[column].head(100)
        for value in sample_data:
            if value is not None:
                max_length = max(max_length, len(str(value)))
        
        # Set width with some padding, but cap at reasonable maximum
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _format_header_row(ws) -> None:
    """Format the header row with bold font and center alignment."""
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:  # First row
        cell.font = header_font
        cell.alignment = header_alignment


def _create_table(ws, df: pd.DataFrame, table_name: str) -> None:
    """Create an Excel table with green styling."""
    # Define table range
    end_column = get_column_letter(len(df.columns))
    end_row = len(df) + 1  # +1 for header
    table_range = f"A1:{end_column}{end_row}"
    
    # Create table
    table = Table(displayName=table_name.replace(" ", "_"), ref=table_range)
    
    # Apply green table style
    style = TableStyleInfo(
        name="TableStyleMedium9",  # Green table style
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    
    # Add table to worksheet
    ws.add_table(table) 